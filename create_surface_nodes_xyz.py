"""
create_surface_nodes_v3.py

Parses the Marc/Mentat mesh file (.dat) and three DVC Tecplot files:
  B0001.dat   -> z-component (w)   "u-component [pixel]" in the original naming
  B0001_X.dat -> x-component (u)   u-component in DVC pixel coords
  B0001_Y.dat -> y-component (v)   v-component in DVC pixel coords

Produces surface_nodes.xlsx with sheets:
  - top_surface_nodes      (node_id, x, y, z)
  - bottom_surface_nodes
  - dvc_nodes              (x, y, z, u, v, w, isValid)

    u  = displacement along DVC x-axis (pixel->mm, then axis-swapped -> Marc y direction)
    v  = displacement along DVC y-axis (pixel->mm, then axis-swapped -> Marc x direction)
    w  = displacement along DVC z-axis (pixel->mm, no axis swap    -> Marc z direction)

Axis-swap convention (same as before, applied to both positions AND displacements):
    DVC file column  ->  Marc frame direction
    px_x (u-comp)   ->  Marc y
    px_y (v-comp)   ->  Marc x
    px_z (w-comp)   ->  Marc z

So after the swap the stored columns in dvc_nodes are:
    u_marc = v_dvc * PIXEL_SIZE   (DVC y-displacement -> Marc x)
    v_marc = u_dvc * PIXEL_SIZE   (DVC x-displacement -> Marc y)
    w_marc = w_dvc * PIXEL_SIZE   (DVC z-displacement -> Marc z)

Pipeline applied to the DVC:
  1. Axis swap: DVC_x -> Marc_y, DVC_y -> Marc_x, DVC_z -> Marc_z
  2. Convert DVC pixels to mm (PIXEL_SIZE).
  3. Translate so that DVC_LANDMARK_PX (a chosen anatomical feature) lands
     exactly on MARC_LANDMARK in the output frame.
  4. (Optional) Apply rotation Rx, Ry, Rz around MARC_LANDMARK.
  5. (Optional) Apply additional fine-tuning translation dx, dy, dz.

The same rotation (step 4) is applied to the displacement vector (u, v, w)
in Marc-frame coords, since a rotation of the coordinate frame rotates
vectors identically. The translation (steps 3 & 5) does NOT affect
displacement vectors (they are free vectors, not position vectors).

Pipeline applied to Marc surface nodes:
  - Just read them straight from the .dat -- no rotation, no translation.

Important: NO per-axis SCALING of DVC positions. The DVC keeps its real
physical dimensions, so its measurement region occupies the correct
sub-volume of the Marc specimen.
"""
import re
import numpy as np
from scipy.spatial.transform import Rotation
from openpyxl import Workbook

# ============================
# EDIT THESE VARIABLES
# ============================
dat_file = r"C:\Users\rahwa.tecle\OneDrive - Imperial College London\FYP\Mesh Only\117matprop.dat"
top_set_name    = "top_surface_nodes"
bottom_set_name = "bottom_surface_nodes"

# Three DVC displacement-component files (same grid, same isValid column)
dvc_w_file = r"C:\Users\rahwa.tecle\OneDrive - Imperial College London\FYP\B0001.dat"    # z-component (w)
dvc_u_file = r"C:\Users\rahwa.tecle\OneDrive - Imperial College London\FYP\B0001_X.dat"  # x-component (u)
dvc_v_file = r"C:\Users\rahwa.tecle\OneDrive - Imperial College London\FYP\B0001_Y.dat"  # y-component (v)

output_xlsx = r"C:\Users\rahwa.tecle\OneDrive - Imperial College London\FYP\Mesh Only\surface_nodes.xlsx"

PIXEL_SIZE = 0.039  # mm/pixel

# ---- Landmark alignment ---- 
# ---- Landmark alignment ---- 
DVC_LANDMARK_PX = np.array([342.0, 304.0, 532.0])    # (px_x, px_y, px_z) in raw DVC pixel coords
MARC_LANDMARK   = np.array([13.41223140, -95.48385620, -1220.89936])  # (x, y, z) in Marc mesh coords (mm)

# Optional rotation around the Marc landmark (degrees), applied to the DVC
# AFTER landmark alignment.
Rx = 0.0
Ry = 0.0
Rz = 0.0

# Optional fine-tuning translation applied to DVC AFTER landmark alignment (mm)
dx = 0.0
dy = 0.0
dz = 0.0
# ============================


# ----------------------------
# Marc .dat parsing
# ----------------------------
FLOAT_TOKEN_RE = re.compile(r"[+-]?(?:\d+\.\d*|\d*\.\d+|\d+)(?:[+-]\d+)")
ID_LINE_RE = re.compile(r"^\s*(\d+)\s*(.*)$")


def parse_marc_float(tok: str) -> float:
    m = re.fullmatch(r"([+-]?(?:\d+\.\d*|\d*\.\d+|\d+))([+-]\d+)", tok.strip())
    if not m:
        raise ValueError(f"Not a Marc float token: {tok!r}")
    return float(m.group(1) + "e" + m.group(2))


def extract_coordinates(filename: str) -> dict:
    coords = {}
    with open(filename, "r", errors="ignore") as f:
        for line in f:
            if line.strip().lower() == "coordinates":
                header = next(f)
                parts = header.split()
                if len(parts) < 2:
                    raise ValueError(f"Unexpected coordinates header: {header!r}")
                npoints = int(parts[1])
                for _ in range(npoints):
                    l = next(f).rstrip("\n")
                    m = ID_LINE_RE.match(l)
                    if not m:
                        raise ValueError(f"Bad coordinate line: {l!r}")
                    nid = int(m.group(1))
                    rest = m.group(2)
                    toks = FLOAT_TOKEN_RE.findall(rest)
                    if len(toks) != 3:
                        raise ValueError(f"Could not parse 3 coords from: {l!r}")
                    x, y, z = (parse_marc_float(t) for t in toks)
                    coords[nid] = (x, y, z)
                break
    if not coords:
        raise ValueError("No `coordinates` section found in file.")
    return coords


def extract_nodes_from_set(filename: str, set_name: str) -> list:
    ids = []
    start_re = re.compile(
        rf"^\s*define\s+\w+\s+set\s+{re.escape(set_name)}\s*$",
        re.IGNORECASE,
    )
    with open(filename, "r", errors="ignore") as f:
        in_set = False
        for line in f:
            if not in_set:
                if start_re.match(line):
                    in_set = True
                continue
            if line.strip() == "":
                break
            low = line.lstrip().lower()
            if low.startswith("define") or low.startswith("end"):
                break
            found = re.findall(r"\d+", line)
            if not found:
                break
            ids.extend(int(x) for x in found)
    if not ids:
        raise ValueError(f"Set {set_name!r} not found or empty.")
    return ids


def extract_tecplot_component(filename: str, col_index: int = 3) -> list:
    """
    Read a Tecplot DVC file and return rows of
    (px_x, px_y, px_z, component_value, isValid).
    col_index: 3 for the displacement column (0-based), 4 for isValid.
    """
    points = []
    with open(filename, "r", errors="ignore") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            up = s.upper()
            if up.startswith(("TITLE", "VARIABLES", "ZONE", "STRANDID", "SOLUTIONTIME")):
                continue
            parts = s.split()
            if len(parts) < 3:
                continue
            try:
                x = float(parts[0])
                y = float(parts[1])
                z = float(parts[2])
                comp  = float(parts[col_index])     if len(parts) > col_index     else None
                valid = int(float(parts[4]))         if len(parts) >= 5            else None
            except ValueError:
                continue
            points.append((x, y, z, comp, valid))
    if not points:
        raise ValueError(f"No numeric point rows found in: {filename}")
    return points


# ----------------------------
# Main
# ----------------------------
def main():
    # ---- Parse Marc mesh ----
    print("Parsing Marc mesh...")
    coords = extract_coordinates(dat_file)
    print(f"  Total nodes: {len(coords)}")

    top_ids    = extract_nodes_from_set(dat_file, top_set_name)
    bottom_ids = extract_nodes_from_set(dat_file, bottom_set_name)
    print(f"  {top_set_name}: {len(top_ids)} nodes")
    print(f"  {bottom_set_name}: {len(bottom_ids)} nodes")

    all_marc_x, all_marc_y, all_marc_z = [], [], []
    surface_data = {}
    for set_name, node_ids in [(top_set_name, top_ids), (bottom_set_name, bottom_ids)]:
        rows = []
        missing = 0
        for nid in node_ids:
            xyz = coords.get(nid)
            if xyz is None:
                missing += 1
                continue
            x, y, z = xyz
            rows.append((nid, x, y, z))
            all_marc_x.append(x); all_marc_y.append(y); all_marc_z.append(z)
        surface_data[set_name] = rows
        if missing:
            print(f"  WARNING: {missing} node IDs in {set_name} not found")

    marc_xmin, marc_xmax = min(all_marc_x), max(all_marc_x)
    marc_ymin, marc_ymax = min(all_marc_y), max(all_marc_y)
    marc_zmin, marc_zmax = min(all_marc_z), max(all_marc_z)
    marc_cx = (marc_xmin + marc_xmax) / 2
    marc_cy = (marc_ymin + marc_ymax) / 2
    marc_cz = (marc_zmin + marc_zmax) / 2

    print(f"\n  Marc surface bbox:")
    print(f"    x: [{marc_xmin:.2f}, {marc_xmax:.2f}]  range {marc_xmax-marc_xmin:.2f}")
    print(f"    y: [{marc_ymin:.2f}, {marc_ymax:.2f}]  range {marc_ymax-marc_ymin:.2f}")
    print(f"    z: [{marc_zmin:.2f}, {marc_zmax:.2f}]  range {marc_zmax-marc_zmin:.2f}")
    print(f"    centre: ({marc_cx:.2f}, {marc_cy:.2f}, {marc_cz:.2f})")

    # ---- Parse DVC -- all three component files ----
    print(f"\nParsing DVC w-component (z): {dvc_w_file}")
    dvc_w_raw = extract_tecplot_component(dvc_w_file, col_index=3)
    print(f"  Raw rows: {len(dvc_w_raw)}")

    print(f"Parsing DVC u-component (x): {dvc_u_file}")
    dvc_u_raw = extract_tecplot_component(dvc_u_file, col_index=3)
    print(f"  Raw rows: {len(dvc_u_raw)}")

    print(f"Parsing DVC v-component (y): {dvc_v_file}")
    dvc_v_raw = extract_tecplot_component(dvc_v_file, col_index=3)
    print(f"  Raw rows: {len(dvc_v_raw)}")

    if not (len(dvc_w_raw) == len(dvc_u_raw) == len(dvc_v_raw)):
        raise ValueError(
            f"DVC files have different row counts: "
            f"w={len(dvc_w_raw)}, u={len(dvc_u_raw)}, v={len(dvc_v_raw)}. "
            f"They must share the same grid."
        )

    # ---- Landmark alignment setup ----
    # DVC landmark: convert raw pixel coords to post-swap, mm coords.
    # Axis swap: (px_x, px_y, px_z) -> (phys_x=py*ps, phys_y=px*ps, phys_z=pz*ps)
    dvc_lm_postswap_mm = np.array([
        DVC_LANDMARK_PX[0] * PIXEL_SIZE,   # was py, now phys_x
        DVC_LANDMARK_PX[1] * PIXEL_SIZE,   # was px, now phys_y
        DVC_LANDMARK_PX[2] * PIXEL_SIZE,   # pz unchanged
    ])
    translation = MARC_LANDMARK - dvc_lm_postswap_mm

    print(f"\n  DVC landmark (pixel):           {DVC_LANDMARK_PX}")
    print(f"  DVC landmark (post-swap, mm):   {dvc_lm_postswap_mm}")
    print(f"  Marc landmark (mm):             {MARC_LANDMARK}")
    print(f"  Landmark-alignment translation: {translation}")
    print(f"  Extra fine-tune translation:    ({dx:+.3f}, {dy:+.3f}, {dz:+.3f}) mm")
    print(f"  DVC rotation about Marc landmark: Rx={Rx:+.3f}  Ry={Ry:+.3f}  Rz={Rz:+.3f} deg")

    rot = Rotation.from_euler("xyz", [Rx, Ry, Rz], degrees=True)
    rot_matrix = rot.as_matrix()  # used to rotate displacement vectors too

    # ---- Apply transform to all DVC points ----
    dvc_aligned = []
    for i, (row_w, row_u, row_v) in enumerate(zip(dvc_w_raw, dvc_u_raw, dvc_v_raw)):
        px, py, pz = row_w[0], row_w[1], row_w[2]
        w_px   = row_w[3]    # z-displacement in DVC pixels
        u_px   = row_u[3]    # x-displacement in DVC pixels
        v_px   = row_v[3]    # y-displacement in DVC pixels
        is_valid = row_w[4]

        # --- POSITION: axis swap + pixel->mm ---
        # DVC (px_x, px_y, px_z) -> Marc (py*ps, px*ps, pz*ps)
        phys_x = px * PIXEL_SIZE
        phys_y = py * PIXEL_SIZE
        phys_z = pz * PIXEL_SIZE

        # Landmark alignment translation
        p = np.array([phys_x, phys_y, phys_z]) + translation
        # Rotation about Marc landmark
        p = rot_matrix @ (p - MARC_LANDMARK) + MARC_LANDMARK
        # Fine-tune translation
        p[0] += dx; p[1] += dy; p[2] += dz

        # --- DISPLACEMENT: axis swap + pixel->mm ---
        # Axis swap mirrors the position swap:
        #   DVC u (along px_x) -> Marc y direction  => store as v_marc
        #   DVC v (along px_y) -> Marc x direction  => store as u_marc
        #   DVC w (along px_z) -> Marc z direction  => store as w_marc
        if None not in (u_px, v_px, w_px):
            u_marc_mm = u_px * PIXEL_SIZE   # DVC y-disp -> Marc x
            v_marc_mm = v_px * PIXEL_SIZE   # DVC x-disp -> Marc y
            w_marc_mm = w_px * PIXEL_SIZE   # DVC z-disp -> Marc z

            # Apply the same rotation that was applied to positions
            # (displacement vectors transform identically to position offsets)
            disp_vec = rot_matrix @ np.array([u_marc_mm, v_marc_mm, w_marc_mm])
            u_out, v_out, w_out = disp_vec[0], disp_vec[1], disp_vec[2]
        else:
            u_out = v_out = w_out = None

        dvc_aligned.append((p[0], p[1], p[2], u_out, v_out, w_out, is_valid))

    valid_aligned = [r for r in dvc_aligned if r[6] == 1]
    print(f"\n  DVC valid points after alignment: {len(valid_aligned)}")
    if valid_aligned:
        vx = np.array([r[0] for r in valid_aligned])
        vy = np.array([r[1] for r in valid_aligned])
        vz = np.array([r[2] for r in valid_aligned])
        vu = np.array([r[3] for r in valid_aligned if r[3] is not None])
        vv = np.array([r[4] for r in valid_aligned if r[4] is not None])
        vw = np.array([r[5] for r in valid_aligned if r[5] is not None])
        print(f"    x: [{vx.min():.2f}, {vx.max():.2f}]")
        print(f"    y: [{vy.min():.2f}, {vy.max():.2f}]")
        print(f"    z: [{vz.min():.2f}, {vz.max():.2f}]")
        if len(vu): print(f"    u (Marc x disp): [{vu.min():.4f}, {vu.max():.4f}] mm")
        if len(vv): print(f"    v (Marc y disp): [{vv.min():.4f}, {vv.max():.4f}] mm")
        if len(vw): print(f"    w (Marc z disp): [{vw.min():.4f}, {vw.max():.4f}] mm")

    print(f"\n  Marc bbox (for reference):")
    print(f"    x: [{marc_xmin:.2f}, {marc_xmax:.2f}]")
    print(f"    y: [{marc_ymin:.2f}, {marc_ymax:.2f}]")
    print(f"    z: [{marc_zmin:.2f}, {marc_zmax:.2f}]")

    # ---- Write Excel ----
    wb = Workbook()
    wb.remove(wb.active)

    for set_name in [top_set_name, bottom_set_name]:
        ws = wb.create_sheet(set_name)
        ws.append(["node_id", "x", "y", "z"])
        for nid, x, y, z in surface_data[set_name]:
            ws.append([nid, x, y, z])

    ws_dvc = wb.create_sheet("dvc_nodes")
    # u = Marc x displacement, v = Marc y displacement, w = Marc z displacement
    ws_dvc.append(["x", "y", "z", "u", "v", "w", "isValid"])
    for ax, ay, az, u_s, v_s, w_s, is_valid in dvc_aligned:
        ws_dvc.append([ax, ay, az, u_s, v_s, w_s, is_valid])

    wb.save(output_xlsx)
    print(f"\nSaved: {output_xlsx}")
    print("Columns in dvc_nodes sheet:")
    print("  x, y, z         = aligned position in Marc frame (mm)")
    print("  u               = displacement along Marc x-axis (mm)")
    print("  v               = displacement along Marc y-axis (mm)")
    print("  w               = displacement along Marc z-axis (mm)")
    print("  isValid         = 1 if DVC correlation was valid, 0 otherwise")


if __name__ == "__main__":
    main()
