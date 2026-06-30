"""
make_bc_proc.py 
Extended to interpolate u, v, w (all three DVC displacement components) and project the full 3D vector onto the loading axis.
surface_nodes.xlsx  ->  apply_bc.proc

Steps:
  1. Read surface_nodes.xlsx -- dvc_nodes sheet now has columns: x, y, z, u, v, w, isValid
       u = displacement along Marc x-axis (mm)
       v = displacement along Marc y-axis (mm)
       w = displacement along Marc z-axis (mm)

  2. Compute a rotated copy of the Marc surface coordinates (rotated about
     INTERP_ROT_PIVOT by INTERP_ROT_EULER).  Rotation is applied to MARC for the
     purpose of interpolation only -- the DVC stays in its original frame so
     its z-layer structure is preserved.

  3. Per-DVC-z-layer 2D interpolation of u, v, w independently onto each rotated Marc surface
     (x, y).

  4. Local-z Gaussian blend across nearby DVC z-layers applied independently to each component.

  5. For each node, project the interpolated displacement vector (u, v, w) onto
     the LOADING_AXIS unit vector n:

         scalar = u*nx + v*ny + w*nz      (dot product / sum of projections)
         dx = scalar * nx
         dy = scalar * ny
         dz = scalar * nz

     This is equivalent to summing the individual axis projections:
         dx = (u*nx)*nx + (v*ny)*nx + (w*nz)*nx  = scalar * nx   etc.

  6. Pair each projected BC vector with its original (un-rotated) Marc node ID
     and write a Mentat .proc file.
"""
import numpy as np
from typing import Optional
from openpyxl import load_workbook
from scipy.interpolate import LinearNDInterpolator, NearestNDInterpolator
from scipy.spatial import cKDTree
from scipy.spatial.transform import Rotation

# ============================
# EDIT THESE PATHS
# ============================
input_xlsx = r"C:\Users\rahwa.tecle\OneDrive - Imperial College London\FYP\Mesh Only\surface_nodes.xlsx"
proc_file  = r"C:\Users\rahwa.tecle\OneDrive - Imperial College London\FYP\Mesh Only\apply_bc.proc"

SURFACE_SHEETS = ("top_surface_nodes", "bottom_surface_nodes")

# ---- Interpolation rotation ----
# These are the rotations found visually with visualise_alignment.py.
# Whatever rotation made the DVC sit nicely inside Marc there, set the same
# Euler angles here.  This script then rotates the MARC nodes by the inverse
# of that, so the DVC keeps its original z-layer structure for interpolation.
#
# INTERP_ROT_PIVOT should match the pivot used in the visualiser (default:
# MARC_LANDMARK from create_surface_nodes_xyz.py).
INTERP_ROT_EULER = (0.0, 22.0, -84.5)
INTERP_ROT_PIVOT = np.array([13.41223140, -95.48385620, -1220.89936])

# ---- Loading axis ----
# The interpolated (u, v, w) displacement vector is projected onto this axis.
# The scalar magnitude (dot product) is then re-applied along the axis to give
# the BC components dx, dy, dz.
#
# "auto" derives the axis from INTERP_ROT_EULER. Override with an explicit [nx, ny, nz] if needed.
LOADING_AXIS = "auto"

# ---- Interpolation parameters ----
SIGMA_Z       = 0.1
SMOOTH_ITERS  = 0
SMOOTH_K      = 6
SMOOTH_ALPHA  = 0.3

# DVC layer detection
Z_LAYER_TOL       = 0.05
MIN_PTS_PER_LAYER = 4

# Mentat output
TABLE_NAME = "linear_ramp_time1"
# ============================


# ---------- helpers ----------

def safe_float(v) -> Optional[float]:
    if v is None: return None
    try: return float(v)
    except Exception: return None


def safe_int(v, default: int = 0) -> int:
    if v is None: return default
    try: return int(float(v))
    except Exception: return default


def read_dvc_nodes_valid(ws) -> np.ndarray:
    """
    Read dvc_nodes sheet. Returns array with columns [x, y, z, u, v, w] for valid rows only.
    """
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(header)}

    # Minimum required columns
    for r in ("x", "y", "z", "isValid"):
        if r not in col:
            raise ValueError(f"dvc_nodes missing column {r!r}. Found: {header}")

    has_uvw = ("u" in col and "v" in col and "w" in col)
    has_w_only = ("w" in col and "u" not in col)

    if not has_uvw and not has_w_only:
        raise ValueError(
            f"dvc_nodes must have either (u, v, w) or at least (w) columns. "
            f"Found: {header}"
        )

    if has_w_only:
        print("  WARNING: dvc_nodes has only 'w' column -- u and v will be 0.0.")
        print("  Re-run create_surface_nodes_v3.py with all three DVC files to fix this.")

    pts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if safe_int(row[col["isValid"]], default=0) != 1:
            continue
        x = safe_float(row[col["x"]])
        y = safe_float(row[col["y"]])
        z = safe_float(row[col["z"]])
        w = safe_float(row[col["w"]])
        u = safe_float(row[col["u"]]) if "u" in col else 0.0
        v = safe_float(row[col["v"]]) if "v" in col else 0.0
        if None in (x, y, z, w):
            continue
        if u is None: u = 0.0
        if v is None: v = 0.0
        pts.append((x, y, z, u, v, w))

    if not pts:
        raise ValueError("No valid (isValid==1) rows in dvc_nodes.")
    return np.array(pts)   # shape (N, 6): x y z u v w


def read_surface_sheet(ws):
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(header)}
    for r in ("node_id", "x", "y", "z"):
        if r not in col:
            raise ValueError(f"sheet missing column {r!r}")
    nids, coords = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        x = safe_float(row[col["x"]]); y = safe_float(row[col["y"]]); z = safe_float(row[col["z"]])
        nid = safe_int(row[col["node_id"]])
        if None in (x, y, z) or nid is None:
            continue
        nids.append(nid)
        coords.append((x, y, z))
    return np.asarray(nids, dtype=int), np.asarray(coords, dtype=float)


def group_into_layers(dvc_pts: np.ndarray, tol: float):
    """Group DVC points by z-layer.  dvc_pts columns: x y z u v w."""
    z_all = dvc_pts[:, 2]
    z_unique_sorted = np.sort(np.unique(np.round(z_all / tol) * tol))
    layers = []
    for z_target in z_unique_sorted:
        mask = np.abs(z_all - z_target) <= tol
        if mask.sum() < MIN_PTS_PER_LAYER:
            continue
        z_centroid = float(z_all[mask].mean())
        # Keep x, y, u, v, w for each point in this layer
        layer_pts = dvc_pts[mask][:, [0, 1, 3, 4, 5]]   # x, y, u, v, w
        layers.append((z_centroid, layer_pts))
    if len(layers) < 2:
        raise ValueError(f"Need >=2 usable DVC layers; found {len(layers)}.")
    layers.sort(key=lambda t: t[0])
    return layers


def build_layer_interpolators(layers):
    """Build per-layer 2D interpolators for u, v, w independently."""
    out = []
    for z_val, xy_uvw in layers:
        xy = xy_uvw[:, :2]        # shape (M, 2)
        u  = xy_uvw[:, 2]
        v  = xy_uvw[:, 3]
        w  = xy_uvw[:, 4]
        interp_u = (LinearNDInterpolator(xy, u), NearestNDInterpolator(xy, u))
        interp_v = (LinearNDInterpolator(xy, v), NearestNDInterpolator(xy, v))
        interp_w = (LinearNDInterpolator(xy, w), NearestNDInterpolator(xy, w))
        out.append((z_val, interp_u, interp_v, interp_w))
    return out


def evaluate_component_at_layers(layer_interps, xy_query: np.ndarray, comp_idx: int):
    """
    Evaluate one displacement component (0=u, 1=v, 2=w) at xy_query for every layer.
    Returns z_vals (L,) and w_mat (N, L).
    comp_idx: 0=u interpolators, 1=v, 2=w  (index into the tuple stored per layer)
    """
    L = len(layer_interps)
    N = len(xy_query)
    z_vals = np.array([z for z, *_ in layer_interps])
    val_mat = np.empty((N, L))
    for j, (_, interp_u, interp_v, interp_w) in enumerate(layer_interps):
        lin, near = (interp_u, interp_v, interp_w)[comp_idx]
        v = lin(xy_query)
        nan_mask = np.isnan(v)
        if nan_mask.any():
            v[nan_mask] = near(xy_query[nan_mask])
        val_mat[:, j] = v
    return z_vals, val_mat

def gaussian_weighted_average(z_layers, w_at_layers, z_query, sigma_z):
    dz = z_query[:, None] - z_layers[None, :]
    weights = np.exp(-(dz / sigma_z) ** 2)
    weight_sum = weights.sum(axis=1, keepdims=True)
    
    # Fallback: where all weights underflowed to zero, use nearest layer
    bad = (weight_sum < 1e-300).flatten()
    if bad.any():
        nearest_layer = np.argmin(np.abs(dz[bad]), axis=1)
        # Use nearest layer's value directly for these queries
        out = np.empty(len(z_query))
        # First do the normal weighted average for good queries
        good_weights = weights[~bad] / weight_sum[~bad]
        out[~bad] = np.sum(good_weights * w_at_layers[~bad], axis=1)
        # Then fill the bad queries with nearest-layer value
        out[bad] = w_at_layers[bad, nearest_layer]
        return out
    
    weights /= weight_sum
    return np.sum(weights * w_at_layers, axis=1)


def smooth_on_surface(coords, values, k, alpha, iters):
    if iters <= 0 or len(coords) < k + 1:
        return values.copy()
    tree = cKDTree(coords)
    dists, idx = tree.query(coords, k=k + 1)
    neigh_idx, neigh_dist = idx[:, 1:], dists[:, 1:]
    w = 1.0 / np.maximum(neigh_dist, 1e-9)
    w /= w.sum(axis=1, keepdims=True)
    v_cur = values.astype(float).copy()
    for _ in range(iters):
        v_cur = (1.0 - alpha) * v_cur + alpha * np.sum(w * v_cur[neigh_idx], axis=1)
    return v_cur


# ---------- main ----------

def main():
    wb = load_workbook(input_xlsx, data_only=True)
    needed = set(SURFACE_SHEETS) | {"dvc_nodes"}
    missing = needed - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Missing sheets: {sorted(missing)}.")

    # ----- DVC -----
    dvc_pts = read_dvc_nodes_valid(wb["dvc_nodes"])
    # Columns: x y z u v w
    print(f"DVC valid points: {len(dvc_pts)}")
    print(f"  x: [{dvc_pts[:,0].min():.2f}, {dvc_pts[:,0].max():.2f}]")
    print(f"  y: [{dvc_pts[:,1].min():.2f}, {dvc_pts[:,1].max():.2f}]")
    print(f"  z: [{dvc_pts[:,2].min():.2f}, {dvc_pts[:,2].max():.2f}]")
    print(f"  u (Marc x-disp): [{dvc_pts[:,3].min():.4f}, {dvc_pts[:,3].max():.4f}] mm")
    print(f"  v (Marc y-disp): [{dvc_pts[:,4].min():.4f}, {dvc_pts[:,4].max():.4f}] mm")
    print(f"  w (Marc z-disp): [{dvc_pts[:,5].min():.4f}, {dvc_pts[:,5].max():.4f}] mm")

    layers = group_into_layers(dvc_pts, tol=Z_LAYER_TOL)
    z_arr  = np.array([L[0] for L in layers])
    print(f"\nUsable DVC z-layers: {len(layers)}")
    print(f"  z: {z_arr[0]:.2f} -> {z_arr[-1]:.2f} mm  "
          f"(median spacing {np.median(np.diff(z_arr)):.2f} mm)")

    print("\nBuilding 2D interpolants per z-layer (u, v, w independently)...")
    layer_interps = build_layer_interpolators(layers)
    print("  done.")

    # ----- Read surface sheets -----
    sheets_data = []
    for sheet_name in SURFACE_SHEETS:
        nids, coords = read_surface_sheet(wb[sheet_name])
        sheets_data.append((sheet_name, nids, coords))
        print(f"\n{sheet_name}: {len(nids)} nodes")
        print(f"  x: [{coords[:,0].min():.2f}, {coords[:,0].max():.2f}]")
        print(f"  y: [{coords[:,1].min():.2f}, {coords[:,1].max():.2f}]")
        print(f"  z: [{coords[:,2].min():.2f}, {coords[:,2].max():.2f}]")

    # ----- Compute the inverse rotation to apply to Marc -----
    R_visualiser = Rotation.from_euler("xyz", INTERP_ROT_EULER, degrees=True)
    R_marc       = R_visualiser.inv()
    print(f"\nVisualiser rotation (applied to DVC during alignment): "
          f"xyz Euler {INTERP_ROT_EULER}")
    print(f"Inverse rotation applied here to MARC nodes about pivot "
          f"{tuple(np.round(INTERP_ROT_PIVOT, 3))}:")
    print(f"  rotation matrix:\n{R_marc.as_matrix()}")

    # ----- Loading axis -----
    def normalise(v):
        v = np.asarray(v, dtype=float)
        n = np.linalg.norm(v)
        if n < 1e-12:
            raise ValueError("zero-length loading axis vector")
        return v / n

    if isinstance(LOADING_AXIS, str) and LOADING_AXIS.lower() == "auto":
        axis = normalise(R_visualiser.apply(np.array([0.0, 0.0, 1.0])))
        print(f"\nLoading axis (auto, from visualiser rotation): "
              f"({axis[0]:+.6f}, {axis[1]:+.6f}, {axis[2]:+.6f})")
        print(f"  angle from file-z: "
              f"{np.degrees(np.arccos(np.clip(abs(axis[2]), -1, 1))):.2f} deg")
    else:
        axis = normalise(LOADING_AXIS)
        print(f"\nLoading axis (explicit): "
              f"({axis[0]:+.6f}, {axis[1]:+.6f}, {axis[2]:+.6f})")

    # ----- Per-surface interpolation -----
    per_node_uvw = {}   # node_id -> (u, v, w)
    duplicate_nodes = 0

    for sheet_name, nids, coords_orig in sheets_data:
        print(f"\n=== Interpolating {sheet_name} ===")

        # Rotate Marc coords for interpolation purposes only
        coords_rot = R_marc.apply(coords_orig - INTERP_ROT_PIVOT) + INTERP_ROT_PIVOT
        print(f"  rotated coords (used for interpolation):")
        print(f"    x: [{coords_rot[:,0].min():.2f}, {coords_rot[:,0].max():.2f}]")
        print(f"    y: [{coords_rot[:,1].min():.2f}, {coords_rot[:,1].max():.2f}]")
        print(f"    z: [{coords_rot[:,2].min():.2f}, {coords_rot[:,2].max():.2f}]")

        xy_rot = coords_rot[:, :2]
        z_rot  = coords_rot[:, 2]

        # Interpolate each component independently
        results = {}
        for comp_name, comp_idx in (("u", 0), ("v", 1), ("w", 2)):
            z_layer_vals, val_at_layers = evaluate_component_at_layers(
                layer_interps, xy_rot, comp_idx
            )
            val_blend = gaussian_weighted_average(z_layer_vals, val_at_layers, z_rot, SIGMA_Z)
            val_final = smooth_on_surface(coords_rot, val_blend,
                                          k=SMOOTH_K, alpha=SMOOTH_ALPHA, iters=SMOOTH_ITERS)
            results[comp_name] = val_final
            print(f"  {comp_name}_interp: min={val_final.min():.4f}  max={val_final.max():.4f}  "
                  f"mean={val_final.mean():.4f}  std={val_final.std():.4f}")

        u_arr = results["u"]
        v_arr = results["v"]
        w_arr = results["w"]

        # Pair with original node IDs
        for nid, u_val, v_val, w_val in zip(nids, u_arr, v_arr, w_arr):
            if nid in per_node_uvw:
                duplicate_nodes += 1
            else:
                per_node_uvw[nid] = (float(u_val), float(v_val), float(w_val))

    if duplicate_nodes:
        print(f"\nNote: {duplicate_nodes} nodes appeared in more than one sheet; first kept.")

    # ----- Write proc file -----
    print(f"\nWriting Mentat proc file: {proc_file}")
    n_total = len(per_node_uvw)
    with open(proc_file, "w", newline="\n", encoding="ascii") as f:
        for i, nid in enumerate(sorted(per_node_uvw.keys())):
            u_interp, v_interp, w_interp = per_node_uvw[nid]
            if i % 5000 == 0:
                print(f"  BC {i+1}/{n_total}...")

            # Rotate the interpolated displacement from DVC frame to Marc frame
            disp_marc = R_visualiser.as_matrix() @ np.array([u_interp, v_interp, w_interp])
            dx, dy, dz = disp_marc[0], disp_marc[1], disp_marc[2]

            f.write("*new_apply\n")
            f.write(f"*apply_name BCR_{nid}\n")
            f.write("*apply_type fixed_displacement\n")
            f.write("*apply_dof x\n")
            f.write(f"*apply_dof_value x {dx:.10e}\n")
            f.write("*apply_dof y\n")
            f.write(f"*apply_dof_value y {dy:.10e}\n")
            f.write("*apply_dof z\n")
            f.write(f"*apply_dof_value z {dz:.10e}\n")
            f.write("@set($dof,x)\n")
            f.write(f"*apply_dof_table x {TABLE_NAME}\n")
            f.write("@set($dof,y)\n")
            f.write(f"*apply_dof_table y {TABLE_NAME}\n")
            f.write("@set($dof,z)\n")
            f.write(f"*apply_dof_table z {TABLE_NAME}\n")
            f.write("*add_apply_nodes\n")
            f.write(f"{nid}\n")
            f.write("#\n")

    print(f"\nDone. Total BCs: {n_total}")
    print(f"Loading axis: ({axis[0]:+.6f}, {axis[1]:+.6f}, {axis[2]:+.6f})")
    print(f"Each BC = dot((u,v,w), axis) * axis  [projection of full 3D vector]")


if __name__ == "__main__":
    main()
